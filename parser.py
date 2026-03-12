"""
D'Casa Honduras ERP report parser.

Reads fixed-width .dat exports, enriches each line item with lookup tables,
and writes the result to an Excel workbook.

Usage:
    python parser.py                        # writes Resultado.xlsx
    python parser.py --out my_output.xlsx   # custom output path
"""

import re
import argparse
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent

SPANISH_MONTHS_ABBR = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}
SPANISH_MONTHS_FULL = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO",
    6: "JUNIO", 7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE",
    10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}
SPANISH_DAYS = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO"]

# Vendor overrides keyed by (ruta_presupuesto, estacion) -> (cod_vnd, vendedor, ruta).
# Used for COD_RUTA==0 clients and for cases where vnd_lookup has no SPS/TEG entry.
SPECIAL_VENDOR = {
    ("HN11", "TEG"): (10, "ETNA VENEGAS WAI",   "Supermercado"),
    ("HN11", "SPS"): (33, "MIRIAM RIVERA",       "Supermercado"),
    ("HN10", "TEG"): ( 6, "ETNA VENEGAS WAI",   "Supermercado"),
    ("HN10", "SPS"): (31, "MIRIAM RIVERA",       "Supermercado"),
    ("HN12", "SPS"): (99, "VENTAS OFICINA SPS",  "Supermercado"),
    ("HN12", "TEG"): ( 7, "VENTAS OFICINA TEG",  "Supermercado"),
    ("HN17", "TEG"): ( 7, "VENTAS OFICINA TEG",  "Oficina"),
    ("HN17", "SPS"): (99, "VENTAS OFICINA SPS",  "Oficina"),
    ("ZCS4", "TEG"): ( 4, "FRANCISCO LOPEZ",     "El Paraiso y alrededores"),
    ("ZCS15","TEG"): ( 9, "ETNA VENEGAS WAI",    "Supermercado"),
    ("ZN24", "SPS"): (24, "INGRID ROSA",         "Occidente 2"),
}

# Patch only the RUTA field when vnd_lookup has the right vendor but wrong category.
RUTA_OVERRIDE = {
    ("ZCS15", "TEG"): "Supermercado",
    ("HN12",  "TEG"): "Supermercado",
}

# Output column order (matches Resultado esperado.xlsx)
COLUMNS = [
    "DOCUMENTO", "LINEA FAC", "BODEGA", "COD PROV", "PROVEEDOR",
    "LINEA", "CODIGO BARRA", "DESCRIPCION", "UNID/VENTA", "CANTIDAD",
    "COSTO", "PRECIO", "DESCUENTO", "ALTERNO", "MODULO", "ESTACION",
    "MES", "AÑO", "FECHA", "DIA", "SEMANA", "CODIGO CLIENTE",
    "DEPARTAMENTO", "MUNICIPIO", "NOMBRE CLIENTE", "TIPO CLIENTE",
    "COD VENDEDOR", "VENDEDOR", "RUTA", "TIPO VENTA", "ORDEN",
    "RAZON", "PESO", "RUTA-PRESUPUESTO",
]

# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def parse_dat_date(s: str) -> datetime:
    """Parse '03.MAR.26' -> datetime(2026, 3, 3)."""
    parts = s.strip().split(".")
    return datetime(2000 + int(parts[2]), SPANISH_MONTHS_ABBR[parts[1]], int(parts[0]))


def semana_del_mes(d: datetime) -> int:
    """Week number within the month (1-based, weeks start on day 1)."""
    return (d.day - 1) // 7 + 1


def _num(s: str) -> float:
    """Parse a numeric string that may contain commas as thousand separators."""
    return float(s.replace(",", "").strip()) if s.strip() else 0.0


# ---------------------------------------------------------------------------
# Parse INVR0601  (lp59.dat = devoluciones,  lp60.dat = ventas)
# ---------------------------------------------------------------------------

def parse_invr0601(filepath: Path, modulo: str, tipo_venta: str) -> list:
    """
    Return a list of dicts, one per invoice line.

    Main line columns (fixed-width):
        DOCUMENTO   [0:19]    LINEA      [20:23]   BODEGA     [26:28]
        COD_PROV    [29:32]   LINEA_PROD [34:36]   COD_BARRA  [38:53]
        DESCRIPCION [54:80]   UNID_VENTA [80:84]   CANTIDAD   [84:95]
        COSTO       [95:108]  PRECIO     [108:121] DESCUENTO  [121:132]

    Alternate line (37 leading spaces): ALTERNO at [37:43]
    """
    records = []
    pending = None

    with open(filepath, encoding="latin-1") as fh:
        for raw in fh:
            line = raw.rstrip()
            if not line:
                continue

            if re.match(r"^\d{3}-\d{3}-\d{2}-\d{8}", line):
                if len(line) < 121:
                    pending = None
                    continue
                pending = {
                    "documento":   line[0:19].strip(),
                    "linea":       int(line[20:23].strip()),
                    "bodega":      int(line[26:28].strip()),
                    "cod_prov":    int(line[29:32].strip()),
                    "linea_prod":  int(line[34:36].strip()),
                    "cod_barra":   line[38:53].strip(),
                    "descripcion": line[54:80].strip(),
                    "unid_venta":  line[80:84].strip(),
                    "cantidad":    int(_num(line[84:95])),
                    "costo":       round(_num(line[95:108]), 2),
                    "precio":      round(_num(line[108:121]), 2),
                    "descuento":   round(_num(line[121:132]), 2) if len(line) > 121 else 0,
                    "alterno":     None,
                    "modulo":      modulo,
                    "tipo_venta":  tipo_venta,
                }
            elif pending is not None:
                if len(line) >= 43 and line[:37] == " " * 37:
                    alt_str = line[37:43].strip()
                    if alt_str.isdigit():
                        pending["alterno"] = int(alt_str)
                        records.append(pending)
                pending = None

    return records


# ---------------------------------------------------------------------------
# Parse VENR15  (lp33.dat)
# ---------------------------------------------------------------------------

_VENR15_RE = re.compile(
    r"^(\d{2}\.\w{3}\.\d{2}) (\d{3}-\d{3}-\d{2}-\d{8}) \d+ \d+ (\d+)"
)

def parse_venr15(filepath: Path) -> dict:
    """Return {documento: {fecha, cliente}}."""
    result = {}
    with open(filepath, encoding="latin-1") as fh:
        for line in fh:
            m = _VENR15_RE.match(line)
            if m:
                result[m.group(2)] = {
                    "fecha":   parse_dat_date(m.group(1)),
                    "cliente": int(m.group(3)),
                }
    return result


# ---------------------------------------------------------------------------
# Parse FACR12  (lp35.dat)
# ---------------------------------------------------------------------------

_FACR12_HEAD = re.compile(r"^(\d{3}-\d{3}-\d{2}-\d{8}) (\d+)")

def parse_facr12(filepath: Path) -> dict:
    """Return {documento: {cliente, orden}}."""
    result = {}
    with open(filepath, encoding="latin-1") as fh:
        for line in fh:
            m = _FACR12_HEAD.match(line)
            if m:
                tokens = line.rstrip().split()
                result[m.group(1)] = {
                    "cliente": int(m.group(2)),
                    "orden":   tokens[-1],
                }
    return result


# ---------------------------------------------------------------------------
# Parse NOTR03  (lp31.dat)
# ---------------------------------------------------------------------------

_NOTR03_RE = re.compile(
    r"^(\d{3}-\d{3}-06-\d{8}) \d+\s+\d+\s+\d+ (\d{2}\.\w{3}\.\d{2}) CT:(\d+)"
)

def parse_notr03(filepath: Path) -> dict:
    """Return {documento: {fecha, cliente}}."""
    result = {}
    with open(filepath, encoding="latin-1") as fh:
        for line in fh:
            m = _NOTR03_RE.match(line)
            if m:
                result[m.group(1)] = {
                    "fecha":   parse_dat_date(m.group(2)),
                    "cliente": int(m.group(3)),
                }
    return result


# ---------------------------------------------------------------------------
# Parse INVR29  (lp51.dat)
# ---------------------------------------------------------------------------

_INVR29_RE = re.compile(
    r"^\s{2}(\d{2}\.\w{3}\.\d{2}) (\d{3}-\d{3}-\d{2}-\d{8})-(\d{3}) (\S.*?)\s{2,}(?:SALIDA|ENTRADA)"
)

def parse_invr29(filepath: Path) -> dict:
    """Return {(documento, linea): razon}."""
    result = {}
    with open(filepath, encoding="latin-1") as fh:
        for line in fh:
            m = _INVR29_RE.match(line.rstrip())
            if m:
                result[(m.group(2), int(m.group(3)))] = m.group(4).strip()
    return result


# ---------------------------------------------------------------------------
# Load Variables.xlsx
# ---------------------------------------------------------------------------

def load_variables(filepath: Path) -> dict:
    """
    Load all lookup tables from Variables.xlsx.

    Returns a dict with keys:
        client_ruta  – {codigo_cliente: {depto, muni, nombre, tipo, cod_ruta, ruta_pres}}
        vnd_lookup   – {(cod_vnd, estacion): {vendedor, ruta}}
        prov_lookup  – {cod_prov: proveedor_name}
        bodega_est   – {bodega_code: estacion}   ('TEG' or 'SPS')
        peso_lookup  – {alterno: peso}
    """
    wb = openpyxl.load_workbook(filepath)

    # ---- RUTA-PRESUPUESTO sheet ----
    ws_ruta = wb["RUTA-PRESUPUESTO"]
    client_ruta = {}
    for row in ws_ruta.iter_rows(values_only=True):
        cli = row[2]
        if cli and isinstance(cli, int):
            client_ruta[cli] = {
                "depto":     row[3],
                "muni":      row[4],
                "nombre":    row[5],
                "tipo":      row[6],
                "cod_ruta":  row[8] if isinstance(row[8], int) else 0,
                "ruta_pres": row[10],
            }

    # ---- Data NUEVO sheet ----
    ws_data = wb["Data NUEVO"]
    vnd_lookup  = {}
    prov_lookup = {}
    bodega_est  = {}

    for row in ws_data.iter_rows(values_only=True):
        cod_vnd  = row[8]
        estacion = row[11]

        if (isinstance(cod_vnd, int) and cod_vnd < 100
                and isinstance(estacion, str) and estacion in ("TEG", "SPS")
                and row[10] and isinstance(row[10], str)):
            key = (cod_vnd, estacion)
            if key not in vnd_lookup:
                vnd_lookup[key] = {"vendedor": row[10], "ruta": row[9]}

        elif isinstance(cod_vnd, int) and cod_vnd >= 100 and row[9] and isinstance(row[9], str):
            prov_lookup[cod_vnd] = row[9]
            if isinstance(row[11], int) and row[12] in ("TEG", "SPS"):
                bodega_est[row[11]] = row[12]

    # ---- Peso PRODUCTOS sheet ----
    ws_peso = wb["Peso PRODUCTOS"]
    peso_lookup = {}
    for row in ws_peso.iter_rows(values_only=True):
        alt  = row[1]
        peso = row[3]
        if alt and isinstance(alt, int) and peso is not None:
            if alt not in peso_lookup:
                peso_lookup[alt] = peso

    return {
        "client_ruta": client_ruta,
        "vnd_lookup":  vnd_lookup,
        "prov_lookup": prov_lookup,
        "bodega_est":  bodega_est,
        "peso_lookup": peso_lookup,
    }


# ---------------------------------------------------------------------------
# Vendor resolution
# ---------------------------------------------------------------------------

def resolve_vendor(client_info: dict, estacion: str, variables: dict) -> tuple:
    """Return (cod_vnd, vendedor, ruta)."""
    cod_ruta  = client_info["cod_ruta"]
    ruta_pres = client_info.get("ruta_pres", "")

    if cod_ruta != 0:
        info = variables["vnd_lookup"].get((cod_ruta, estacion))
        if info:
            ruta = RUTA_OVERRIDE.get((ruta_pres, estacion), info.get("ruta", ""))
            return cod_ruta, info.get("vendedor", ""), ruta
        key = (ruta_pres, estacion)
        if key in SPECIAL_VENDOR:
            return SPECIAL_VENDOR[key]
        return cod_ruta, "", ""

    key = (ruta_pres, estacion)
    if key in SPECIAL_VENDOR:
        return SPECIAL_VENDOR[key]
    return 0, "", ""


# ---------------------------------------------------------------------------
# Build the full output dataset
# ---------------------------------------------------------------------------

def build_dataset(dat_dir: Path = None) -> list:
    """
    Parse all source files and return a list of enriched row dicts.

    dat_dir  – directory containing the .dat files (defaults to BASE_DIR).
               Variables.xlsx is always loaded from BASE_DIR.
    """
    if dat_dir is None:
        dat_dir = BASE_DIR

    variables  = load_variables(BASE_DIR / "Variables.xlsx")
    sales      = parse_invr0601(dat_dir / "lp60.dat", "FACTURACION",  "VENTA")
    returns    = parse_invr0601(dat_dir / "lp59.dat", "DEVOLUCION",   "NOTA DE CREDITO")
    doc_venr15 = parse_venr15(dat_dir / "lp33.dat")
    doc_facr12 = parse_facr12(dat_dir / "lp35.dat")
    doc_notr03 = parse_notr03(dat_dir / "lp31.dat")
    razon_map  = parse_invr29(dat_dir / "lp51.dat")

    output = []

    for item in sales + returns:
        doc    = item["documento"]
        linea  = item["linea"]
        bodega = item["bodega"]

        if doc in doc_venr15:
            fecha          = doc_venr15[doc]["fecha"]
            codigo_cliente = doc_venr15[doc]["cliente"]
        elif doc in doc_notr03:
            fecha          = doc_notr03[doc]["fecha"]
            codigo_cliente = doc_notr03[doc]["cliente"]
        else:
            continue

        orden        = doc_facr12[doc]["orden"] if doc in doc_facr12 else 0
        razon        = razon_map.get((doc, linea), "")
        estacion     = variables["bodega_est"].get(bodega, "TEG")
        client_info  = variables["client_ruta"].get(codigo_cliente)
        if client_info is None:
            continue

        cod_vnd, vendedor, ruta = resolve_vendor(client_info, estacion, variables)
        proveedor    = variables["prov_lookup"].get(item["cod_prov"], "")

        tipo_cliente = client_info["tipo"]
        ruta_pres    = client_info["ruta_pres"]

        descuento = item["descuento"]
        if tipo_cliente in ("PROVEEDOR ", "OFICINA ") or ruta_pres == "HN10":
            descuento = 0.0

        depto = client_info["depto"]
        muni  = client_info["muni"]
        if ruta_pres == "HN11" and estacion == "SPS":
            depto = "CORTES"
            muni  = "SAN PEDRO SULA"

        if bodega in (3, 4) and item["alterno"]:
            peso = round(variables["peso_lookup"].get(item["alterno"], 0) * item["cantidad"], 6)
        else:
            peso = 0

        output.append({
            "DOCUMENTO":        doc,
            "LINEA FAC":        linea,
            "BODEGA":           bodega,
            "COD PROV":         item["cod_prov"],
            "PROVEEDOR":        proveedor,
            "LINEA":            item["linea_prod"],
            "CODIGO BARRA":     item["cod_barra"],
            "DESCRIPCION":      item["descripcion"],
            "UNID/VENTA":       item["unid_venta"],
            "CANTIDAD":         item["cantidad"],
            "COSTO":            item["costo"],
            "PRECIO":           item["precio"],
            "DESCUENTO":        descuento,
            "ALTERNO":          item["alterno"],
            "MODULO":           item["modulo"],
            "ESTACION":         estacion,
            "MES":              SPANISH_MONTHS_FULL[fecha.month],
            "AÑO":              fecha.year,
            "FECHA":            fecha,
            "DIA":              SPANISH_DAYS[fecha.weekday()],
            "SEMANA":           semana_del_mes(fecha),
            "CODIGO CLIENTE":   codigo_cliente,
            "DEPARTAMENTO":     depto,
            "MUNICIPIO":        muni,
            "NOMBRE CLIENTE":   client_info["nombre"],
            "TIPO CLIENTE":     tipo_cliente,
            "COD VENDEDOR":     cod_vnd,
            "VENDEDOR":         vendedor,
            "RUTA":             ruta,
            "TIPO VENTA":       item["tipo_venta"],
            "ORDEN":            orden,
            "RAZON":            razon,
            "PESO":             peso,
            "RUTA-PRESUPUESTO": ruta_pres,
        })

    return output


# ---------------------------------------------------------------------------
# Write Excel output
# ---------------------------------------------------------------------------

def write_excel(rows: list, out_path: Path) -> None:
    """Write the dataset to an Excel workbook with a styled header row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MAIN"

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

    # Column widths (rough auto-fit)
    COL_WIDTHS = {
        "DOCUMENTO": 22, "LINEA FAC": 9, "BODEGA": 8, "COD PROV": 9,
        "PROVEEDOR": 22, "LINEA": 7, "CODIGO BARRA": 17, "DESCRIPCION": 28,
        "UNID/VENTA": 10, "CANTIDAD": 10, "COSTO": 14, "PRECIO": 14,
        "DESCUENTO": 12, "ALTERNO": 9, "MODULO": 14, "ESTACION": 9,
        "MES": 12, "AÑO": 7, "FECHA": 14, "DIA": 11, "SEMANA": 8,
        "CODIGO CLIENTE": 15, "DEPARTAMENTO": 22, "MUNICIPIO": 22,
        "NOMBRE CLIENTE": 28, "TIPO CLIENTE": 18, "COD VENDEDOR": 12,
        "VENDEDOR": 22, "RUTA": 22, "TIPO VENTA": 16, "ORDEN": 12,
        "RAZON": 24, "PESO": 10, "RUTA-PRESUPUESTO": 18,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 14)

    wb.save(out_path)
    print(f"Wrote {len(rows):,} rows to {out_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="D'Casa Honduras ERP parser")
    ap.add_argument("--dir", default=None,
                    help="Directory containing the .dat files (default: script directory)")
    ap.add_argument("--out", default="Resultado.xlsx",
                    help="Output Excel file (default: Resultado.xlsx)")
    args = ap.parse_args()

    dat_dir = Path(args.dir) if args.dir else None

    print("Parsing source files...")
    rows = build_dataset(dat_dir)
    print(f"Built {len(rows):,} rows.")

    out_path = Path(args.out) if Path(args.out).is_absolute() else BASE_DIR / args.out
    write_excel(rows, out_path)


if __name__ == "__main__":
    main()
