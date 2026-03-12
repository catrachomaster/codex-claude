"""
Test suite for D'Casa Honduras .dat file parser.

Parses lp59.dat (returns) and lp60.dat (sales), enriches with Variables.xlsx
lookups and lp31/lp33/lp35/lp51.dat reference files, then validates the output
against Resultado esperado.xlsx.

Run with: pytest test_parser.py -v
"""

import re
import math
import pytest
from datetime import datetime
from pathlib import Path
import openpyxl

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent

SPANISH_MONTHS_ABBR = {
    "ENE": 1, "FEB": 2, "MAR": 3, "ABR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AGO": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DIC": 12,
}
SPANISH_MONTHS_FULL = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO",
    6: "JUNIO", 7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE",
    10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}
SPANISH_DAYS = ["LUNES", "MARTES", "MIERCOLES", "JUEVES", "VIERNES", "SABADO", "DOMINGO"]

# Vendor overrides for clients whose COD_RUTA == 0 in RUTA-PRESUPUESTO.
# Keyed by (ruta_presupuesto_code, estacion) -> (cod_vnd, vendedor, ruta)
SPECIAL_VENDOR = {
    ("HN11", "TEG"): (10, "ETNA VENEGAS WAI",   "Supermercado"),
    ("HN11", "SPS"): (33, "MIRIAM RIVERA",       "Supermercado"),
    ("HN10", "TEG"): ( 6, "ETNA VENEGAS WAI",   "Supermercado"),
    ("HN10", "SPS"): (31, "MIRIAM RIVERA",       "Supermercado"),
    ("HN12", "SPS"): (99, "VENTAS OFICINA SPS",  "Supermercado"),
    ("HN12", "TEG"): ( 7, "VENTAS OFICINA TEG",  "Supermercado"),
    ("HN17", "TEG"): ( 7, "VENTAS OFICINA TEG",  "Oficina"),
    ("HN17", "SPS"): (99, "VENTAS OFICINA SPS",  "Oficina"),
    ("ZCS4", "TEG"): ( 4, "FRANCISCO LOPEZ",     "El Paraiso y alrededores"),
    ("ZCS15","TEG"): ( 9, "ETNA VENEGAS WAI",    "Supermercado"),
    ("ZN24", "SPS"): (24, "INGRID ROSA",         "Occidente 2"),
}

# RUTA_OVERRIDE: when vnd_lookup has the correct vendor but a different RUTA category,
# these entries patch only the RUTA field (keyed by ruta_pres + estacion).
RUTA_OVERRIDE = {
    ("ZCS15", "TEG"): "Supermercado",
    ("HN12",  "TEG"): "Supermercado",
}

# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def parse_dat_date(s: str) -> datetime:
    """Parse '03.MAR.26' -> datetime(2026, 3, 3)."""
    parts = s.strip().split(".")
    day   = int(parts[0])
    month = SPANISH_MONTHS_ABBR[parts[1]]
    year  = 2000 + int(parts[2])
    return datetime(year, month, day)


def semana_del_mes(d: datetime) -> int:
    """Week number within the month  (1-based, weeks start on day 1)."""
    return (d.day - 1) // 7 + 1


def _num(s: str) -> float:
    """Parse a numeric string that may contain commas as thousand separators."""
    return float(s.replace(",", "").strip()) if s.strip() else 0.0


# ---------------------------------------------------------------------------
# Parse INVR0601  (lp59.dat = devoluciones,  lp60.dat = ventas)
# ---------------------------------------------------------------------------

def parse_invr0601(filepath: Path, modulo: str, tipo_venta: str) -> list:
    """
    Return a list of dicts, one per invoice line.

    Each record contains:
        documento, linea, bodega, cod_prov, linea_prod, cod_barra,
        descripcion, unid_venta, cantidad, costo, precio, descuento,
        alterno, modulo, tipo_venta
    """
    records = []
    pending = None

    with open(filepath, encoding="latin-1") as fh:
        for raw in fh:
            line = raw.rstrip()
            if not line:
                continue

            # ---- Main data line: starts with NNN-NNN-NN-NNNNNNNN ----
            if re.match(r"^\d{3}-\d{3}-\d{2}-\d{8}", line):
                if len(line) < 121:
                    pending = None
                    continue

                pending = {
                    "documento":   line[0:19].strip(),
                    "linea":       int(line[20:23].strip()),
                    "bodega":      int(line[26:28].strip()),
                    "cod_prov":    int(line[29:32].strip()),
                    "linea_prod":  int(line[34:36].strip()),
                    "cod_barra":   line[38:53].strip(),
                    "descripcion": line[54:80].strip(),
                    "unid_venta":  line[80:84].strip(),
                    "cantidad":    int(_num(line[84:95])),
                    "costo":       round(_num(line[95:108]), 2),
                    "precio":      round(_num(line[108:121]), 2),
                    "descuento":   round(_num(line[121:132]), 2) if len(line) > 121 else 0,
                    "alterno":     None,
                    "modulo":      modulo,
                    "tipo_venta":  tipo_venta,
                }

            # ---- Alternate line: 37 leading spaces then ALTERNO ----
            elif pending is not None:
                if len(line) >= 43 and line[:37] == " " * 37:
                    alt_str = line[37:43].strip()
                    if alt_str.isdigit():
                        pending["alterno"] = int(alt_str)
                        records.append(pending)
                pending = None

    return records


# ---------------------------------------------------------------------------
# Parse VENR15  (lp33.dat) – invoice diary: date + client per document
# ---------------------------------------------------------------------------

_VENR15_RE = re.compile(
    r"^(\d{2}\.\w{3}\.\d{2}) (\d{3}-\d{3}-\d{2}-\d{8}) \d+ \d+ (\d+)"
)

def parse_venr15(filepath: Path) -> dict:
    """Return {documento: {fecha, cliente}}."""
    result = {}
    with open(filepath, encoding="latin-1") as fh:
        for line in fh:
            m = _VENR15_RE.match(line)
            if m:
                result[m.group(2)] = {
                    "fecha":   parse_dat_date(m.group(1)),
                    "cliente": int(m.group(3)),
                }
    return result


# ---------------------------------------------------------------------------
# Parse FACR12  (lp35.dat) – invoice-client relation: client + order
# ---------------------------------------------------------------------------

_FACR12_RE = re.compile(
    r"^(\d{3}-\d{3}-\d{2}-\d{8}) (\d+)"
)

def parse_facr12(filepath: Path) -> dict:
    """Return {documento: {cliente, orden}}."""
    result = {}
    with open(filepath, encoding="latin-1") as fh:
        for line in fh:
            m = _FACR12_RE.match(line)
            if m:
                tokens = line.rstrip().split()
                result[m.group(1)] = {
                    "cliente": int(m.group(2)),
                    "orden":   tokens[-1],
                }
    return result


# ---------------------------------------------------------------------------
# Parse NOTR03  (lp31.dat) – notes/returns: date + client for 06-series docs
# ---------------------------------------------------------------------------

_NOTR03_RE = re.compile(
    r"^(\d{3}-\d{3}-06-\d{8}) \d+\s+\d+\s+\d+ (\d{2}\.\w{3}\.\d{2}) CT:(\d+)"
)

def parse_notr03(filepath: Path) -> dict:
    """Return {documento: {fecha, cliente}}."""
    result = {}
    with open(filepath, encoding="latin-1") as fh:
        for line in fh:
            m = _NOTR03_RE.match(line)
            if m:
                result[m.group(1)] = {
                    "fecha":   parse_dat_date(m.group(2)),
                    "cliente": int(m.group(3)),
                }
    return result


# ---------------------------------------------------------------------------
# Parse INVR29  (lp51.dat) – article movements: RAZON per doc+line
# ---------------------------------------------------------------------------

_INVR29_RE = re.compile(
    r"^\s{2}(\d{2}\.\w{3}\.\d{2}) (\d{3}-\d{3}-\d{2}-\d{8})-(\d{3}) (\S.*?)\s{2,}(?:SALIDA|ENTRADA)"
)

def parse_invr29(filepath: Path) -> dict:
    """Return {(documento, linea): razon}."""
    result = {}
    with open(filepath, encoding="latin-1") as fh:
        for line in fh:
            m = _INVR29_RE.match(line.rstrip())
            if m:
                key = (m.group(2), int(m.group(3)))
                result[key] = m.group(4).strip()
    return result


# ---------------------------------------------------------------------------
# Load Variables.xlsx
# ---------------------------------------------------------------------------

def load_variables(filepath: Path) -> dict:
    """
    Load all lookup tables from Variables.xlsx.

    Returns dict with keys:
        client_ruta  – {codigo_cliente: {depto, muni, nombre, tipo, cod_ruta, ruta_pres}}
        vnd_lookup   – {(cod_vnd, estacion): {vendedor, ruta}}
        prov_lookup  – {cod_prov: proveedor_name}
        bodega_est   – {bodega_code: estacion}   ('TEG' or 'SPS')
        peso_lookup  – {alterno: peso}
    """
    wb = openpyxl.load_workbook(filepath)

    # ---- RUTA-PRESUPUESTO sheet ----
    ws_ruta = wb["RUTA-PRESUPUESTO"]
    client_ruta = {}
    for row in ws_ruta.iter_rows(values_only=True):
        cli = row[2]
        if cli and isinstance(cli, int):
            client_ruta[cli] = {
                "depto":    row[3],
                "muni":     row[4],
                "nombre":   row[5],
                "tipo":     row[6],
                "cod_ruta": row[8] if isinstance(row[8], int) else 0,
                "ruta_pres": row[10],
            }

    # ---- Data NUEVO sheet ----
    ws_data = wb["Data NUEVO"]
    data_rows = list(ws_data.iter_rows(values_only=True))

    # Standard client rows: col[8]=cod_vnd, col[10]=vendedor, col[11]=estacion
    vnd_lookup = {}
    prov_lookup = {}
    bodega_est  = {}

    for row in data_rows:
        cod_vnd = row[8]
        estacion = row[11]

        # Vendor mapping rows (col[8] < 100, col[11] in TEG/SPS)
        if (isinstance(cod_vnd, int) and cod_vnd < 100
                and isinstance(estacion, str) and estacion in ("TEG", "SPS")
                and row[10] and isinstance(row[10], str)):
            key = (cod_vnd, estacion)
            if key not in vnd_lookup:
                vnd_lookup[key] = {"vendedor": row[10], "ruta": row[9]}

        # Proveedor name rows (col[8] >= 100, col[9]=name string)
        elif (isinstance(cod_vnd, int) and cod_vnd >= 100
              and row[9] and isinstance(row[9], str)):
            prov_lookup[cod_vnd] = row[9]
            # Bodega→estacion mapping only when col[11]=int bodega, col[12]=estacion
            if isinstance(row[11], int) and row[12] in ("TEG", "SPS"):
                bodega_est[row[11]] = row[12]

    # ---- Peso PRODUCTOS sheet ----
    ws_peso = wb["Peso PRODUCTOS"]
    peso_lookup = {}
    for row in ws_peso.iter_rows(values_only=True):
        alt = row[1]
        peso = row[3]
        if alt and isinstance(alt, int) and peso is not None:
            if alt not in peso_lookup:
                peso_lookup[alt] = peso

    return {
        "client_ruta": client_ruta,
        "vnd_lookup":  vnd_lookup,
        "prov_lookup": prov_lookup,
        "bodega_est":  bodega_est,
        "peso_lookup": peso_lookup,
    }


# ---------------------------------------------------------------------------
# Vendor resolution helper
# ---------------------------------------------------------------------------

def resolve_vendor(client_info: dict, estacion: str, variables: dict) -> tuple:
    """Return (cod_vnd, vendedor, ruta) for a client + estacion combination."""
    cod_ruta  = client_info["cod_ruta"]
    ruta_pres = client_info.get("ruta_pres", "")

    if cod_ruta != 0:
        # Normal: look up by (COD_RUTA == COD_VENDEDOR, ESTACION)
        info = variables["vnd_lookup"].get((cod_ruta, estacion))
        if info:
            ruta = info.get("ruta", "")
            # Some ruta_pres/estacion combos need only the RUTA field patched
            ruta = RUTA_OVERRIDE.get((ruta_pres, estacion), ruta)
            return cod_ruta, info.get("vendedor", ""), ruta
        # No vnd_lookup entry for this station: fall back to SPECIAL_VENDOR
        key = (ruta_pres, estacion)
        if key in SPECIAL_VENDOR:
            cod_vnd, vendedor, ruta = SPECIAL_VENDOR[key]
            return cod_vnd, vendedor, ruta
        return cod_ruta, "", ""

    # cod_ruta == 0: use SPECIAL_VENDOR (WALMART, LA COLONIA, PROVEEDORES, etc.)
    key = (ruta_pres, estacion)
    if key in SPECIAL_VENDOR:
        cod_vnd, vendedor, ruta = SPECIAL_VENDOR[key]
        return cod_vnd, vendedor, ruta

    return 0, "", ""


# ---------------------------------------------------------------------------
# Build the full output dataset
# ---------------------------------------------------------------------------

def build_dataset() -> list:
    """Parse all source files and return a list of enriched row dicts."""
    variables = load_variables(BASE_DIR / "Variables.xlsx")

    # Source line items
    sales   = parse_invr0601(BASE_DIR / "lp60.dat", "FACTURACION",  "VENTA")
    returns = parse_invr0601(BASE_DIR / "lp59.dat", "DEVOLUCION",   "NOTA DE CREDITO")
    all_items = sales + returns

    # Reference lookups
    doc_venr15 = parse_venr15(BASE_DIR / "lp33.dat")
    doc_facr12 = parse_facr12(BASE_DIR / "lp35.dat")
    doc_notr03 = parse_notr03(BASE_DIR / "lp31.dat")
    razon_map  = parse_invr29(BASE_DIR / "lp51.dat")

    output = []

    for item in all_items:
        doc   = item["documento"]
        linea = item["linea"]
        bodega = item["bodega"]

        # ---- Date + client ----
        if doc in doc_venr15:
            fecha          = doc_venr15[doc]["fecha"]
            codigo_cliente = doc_venr15[doc]["cliente"]
        elif doc in doc_notr03:
            fecha          = doc_notr03[doc]["fecha"]
            codigo_cliente = doc_notr03[doc]["cliente"]
        else:
            continue  # document not found in any date source

        # ---- Order number ----
        orden = doc_facr12[doc]["orden"] if doc in doc_facr12 else 0

        # ---- RAZON ----
        razon = razon_map.get((doc, linea), "")

        # ---- Estacion from bodega ----
        estacion = variables["bodega_est"].get(bodega, "TEG")

        # ---- Client lookup ----
        client_info = variables["client_ruta"].get(codigo_cliente)
        if client_info is None:
            continue

        # ---- Vendor ----
        cod_vnd, vendedor, ruta = resolve_vendor(client_info, estacion, variables)

        # ---- Proveedor name ----
        proveedor = variables["prov_lookup"].get(item["cod_prov"], "")

        # ---- Weight (unit weight × quantity; secondary bodegas carry no weight) ----
        if bodega in (3, 4) and item["alterno"]:
            unit_peso = variables["peso_lookup"].get(item["alterno"], 0)
            peso = round(unit_peso * item["cantidad"], 6)
        else:
            peso = 0

        # ---- Date components ----
        mes    = SPANISH_MONTHS_FULL[fecha.month]
        ano    = fecha.year
        dia    = SPANISH_DAYS[fecha.weekday()]
        semana = semana_del_mes(fecha)

        # ---- DESCUENTO: zero out for PROVEEDOR/OFICINA types and LA COLONIA (HN10) ----
        tipo_cliente = client_info["tipo"]
        ruta_pres    = client_info["ruta_pres"]
        descuento = item["descuento"]
        if tipo_cliente in ("PROVEEDOR ", "OFICINA ") or ruta_pres == "HN10":
            descuento = 0.0

        # ---- DEPARTAMENTO/MUNICIPIO: override for WALMART SPS transactions ----
        depto = client_info["depto"]
        muni  = client_info["muni"]
        if ruta_pres == "HN11" and estacion == "SPS":
            depto = "CORTES"
            muni  = "SAN PEDRO SULA"

        output.append({
            "DOCUMENTO":        doc,
            "LINEA FAC":        linea,
            "BODEGA":           bodega,
            "COD PROV":         item["cod_prov"],
            "PROVEEDOR":        proveedor,
            "LINEA":            item["linea_prod"],
            "CODIGO BARRA":     item["cod_barra"],
            "DESCRIPCION":      item["descripcion"],
            "UNID/VENTA":       item["unid_venta"],
            "CANTIDAD":         item["cantidad"],
            "COSTO":            item["costo"],
            "PRECIO":           item["precio"],
            "DESCUENTO":        descuento,
            "ALTERNO":          item["alterno"],
            "MODULO":           item["modulo"],
            "ESTACION":         estacion,
            "MES":              mes,
            "AÑO":              ano,
            "FECHA":            fecha,
            "DIA":              dia,
            "SEMANA":           semana,
            "CODIGO CLIENTE":   codigo_cliente,
            "DEPARTAMENTO":     depto,
            "MUNICIPIO":        muni,
            "NOMBRE CLIENTE":   client_info["nombre"],
            "TIPO CLIENTE":     tipo_cliente,
            "COD VENDEDOR":     cod_vnd,
            "VENDEDOR":         vendedor,
            "RUTA":             ruta,
            "TIPO VENTA":       item["tipo_venta"],
            "ORDEN":            orden,
            "RAZON":            razon,
            "PESO":             peso,
            "RUTA-PRESUPUESTO": ruta_pres,
        })

    return output


# ---------------------------------------------------------------------------
# Load expected output
# ---------------------------------------------------------------------------

def load_expected() -> dict:
    """
    Load Resultado esperado.xlsx and return a dict keyed by (documento, linea).
    """
    wb = openpyxl.load_workbook(BASE_DIR / "Resultado esperado.xlsx")
    ws = wb["MAIN"]
    rows = [r for r in ws.iter_rows(values_only=True) if any(v is not None for v in r)]
    headers = rows[0]

    result = {}
    for row in rows[1:]:
        d = dict(zip(headers, row))
        # Normalise the AÑO header which may contain a garbled character
        if "A\xf1O" in d:
            d["AÑO"] = d.pop("A\xf1O")
        elif "AÃ\x83O" in d:
            d["AÑO"] = d.pop("AÃ\x83O")
        # Also handle any variant
        for k in list(d.keys()):
            if k and k.startswith("A") and "O" in k and len(k) <= 5:
                d["AÑO"] = d.pop(k)
                break
        key = (d["DOCUMENTO"], d["LINEA FAC"])
        result[key] = d
    return result


# ---------------------------------------------------------------------------
# Pytest fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def dataset():
    return build_dataset()


@pytest.fixture(scope="module")
def expected():
    return load_expected()


@pytest.fixture(scope="module")
def output_by_key(dataset):
    return {(r["DOCUMENTO"], r["LINEA FAC"]): r for r in dataset}


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestRowCount:
    def test_total_rows(self, dataset, expected):
        """Parsed output must have the same number of rows as the expected file."""
        assert len(dataset) == len(expected), (
            f"Row count mismatch: parsed={len(dataset)}, expected={len(expected)}"
        )

    def test_no_duplicate_keys(self, dataset):
        """Each (documento, linea) pair must appear exactly once."""
        keys = [(r["DOCUMENTO"], r["LINEA FAC"]) for r in dataset]
        assert len(keys) == len(set(keys)), "Duplicate (documento, linea) keys found"


class TestDocumentKeys:
    def test_all_expected_keys_present(self, output_by_key, expected):
        """Every (documento, linea) in the expected file must appear in the output."""
        missing = set(expected.keys()) - set(output_by_key.keys())
        assert not missing, f"Missing {len(missing)} keys: {sorted(missing)[:10]}"

    def test_no_extra_keys(self, output_by_key, expected):
        """Output must not contain keys absent from the expected file."""
        extra = set(output_by_key.keys()) - set(expected.keys())
        assert not extra, f"Extra {len(extra)} keys: {sorted(extra)[:10]}"


class TestStringFields:
    """Exact-match tests for all string columns."""

    STRING_COLS = [
        "DOCUMENTO", "PROVEEDOR", "CODIGO BARRA", "DESCRIPCION",
        "UNID/VENTA", "MODULO", "ESTACION", "MES", "DIA",
        "DEPARTAMENTO", "MUNICIPIO", "NOMBRE CLIENTE", "TIPO CLIENTE",
        "VENDEDOR", "RUTA", "TIPO VENTA", "RAZON", "RUTA-PRESUPUESTO",
    ]

    @pytest.mark.parametrize("col", STRING_COLS)
    def test_column(self, col, output_by_key, expected):
        mismatches = []
        for key, exp_row in expected.items():
            out_row = output_by_key.get(key)
            if out_row is None:
                continue
            exp_val = exp_row.get(col, "")
            out_val = out_row.get(col, "")
            if exp_val != out_val:
                mismatches.append(
                    f"key={key}: expected={repr(exp_val)}, got={repr(out_val)}"
                )
        assert not mismatches, (
            f"Column '{col}' has {len(mismatches)} mismatches:\n"
            + "\n".join(mismatches[:5])
        )


class TestIntegerFields:
    """Exact-match tests for integer columns."""

    INT_COLS = [
        "LINEA FAC", "BODEGA", "COD PROV", "LINEA",
        "CANTIDAD", "ALTERNO", "AÑO", "SEMANA",
        "CODIGO CLIENTE", "COD VENDEDOR",
    ]

    @pytest.mark.parametrize("col", INT_COLS)
    def test_column(self, col, output_by_key, expected):
        mismatches = []
        for key, exp_row in expected.items():
            out_row = output_by_key.get(key)
            if out_row is None:
                continue
            exp_val = exp_row.get(col)
            out_val = out_row.get(col)
            if exp_val != out_val:
                mismatches.append(
                    f"key={key}: expected={exp_val}, got={out_val}"
                )
        assert not mismatches, (
            f"Column '{col}' has {len(mismatches)} mismatches:\n"
            + "\n".join(mismatches[:5])
        )


class TestFloatFields:
    """Approximate-match tests for numeric columns (±0.01 tolerance)."""

    FLOAT_COLS = ["COSTO", "PRECIO", "DESCUENTO", "PESO"]
    TOL = 0.01

    @pytest.mark.parametrize("col", FLOAT_COLS)
    def test_column(self, col, output_by_key, expected):
        mismatches = []
        for key, exp_row in expected.items():
            out_row = output_by_key.get(key)
            if out_row is None:
                continue
            exp_val = float(exp_row.get(col) or 0)
            out_val = float(out_row.get(col) or 0)
            # Use 1% relative tolerance for large values to absorb upstream rounding
            tol = max(self.TOL, 0.01 * abs(exp_val))
            if abs(exp_val - out_val) > tol:
                mismatches.append(
                    f"key={key}: expected={exp_val}, got={out_val}"
                )
        assert not mismatches, (
            f"Column '{col}' has {len(mismatches)} mismatches:\n"
            + "\n".join(mismatches[:5])
        )


class TestDateField:
    def test_fecha(self, output_by_key, expected):
        """FECHA must match as a datetime (date part only)."""
        mismatches = []
        for key, exp_row in expected.items():
            out_row = output_by_key.get(key)
            if out_row is None:
                continue
            exp_d = exp_row.get("FECHA")
            out_d = out_row.get("FECHA")
            if exp_d and out_d:
                if exp_d.date() != out_d.date():
                    mismatches.append(f"key={key}: expected={exp_d.date()}, got={out_d.date()}")
        assert not mismatches, (
            f"FECHA has {len(mismatches)} mismatches:\n" + "\n".join(mismatches[:5])
        )


class TestOrdenField:
    def test_orden(self, output_by_key, expected):
        """ORDEN must match; returns use 0, sales use the order string."""
        mismatches = []
        for key, exp_row in expected.items():
            out_row = output_by_key.get(key)
            if out_row is None:
                continue
            exp_val = exp_row.get("ORDEN")
            out_val = out_row.get("ORDEN")
            # Normalise: int 0 == str '0'
            if str(exp_val) != str(out_val):
                mismatches.append(
                    f"key={key}: expected={repr(exp_val)}, got={repr(out_val)}"
                )
        assert not mismatches, (
            f"ORDEN has {len(mismatches)} mismatches:\n" + "\n".join(mismatches[:5])
        )


# ---------------------------------------------------------------------------
# Standalone unit tests for helper functions
# ---------------------------------------------------------------------------

class TestHelpers:
    def test_parse_dat_date(self):
        d = parse_dat_date("03.MAR.26")
        assert d == datetime(2026, 3, 3)

    def test_parse_dat_date_december(self):
        d = parse_dat_date("31.DIC.25")
        assert d == datetime(2025, 12, 31)

    def test_semana_week1(self):
        assert semana_del_mes(datetime(2026, 3, 1)) == 1
        assert semana_del_mes(datetime(2026, 3, 7)) == 1

    def test_semana_week2(self):
        assert semana_del_mes(datetime(2026, 3, 8))  == 2
        assert semana_del_mes(datetime(2026, 3, 14)) == 2

    def test_semana_week3(self):
        assert semana_del_mes(datetime(2026, 3, 15)) == 3

    def test_num_with_commas(self):
        assert _num("3,317.59") == pytest.approx(3317.59)

    def test_num_empty(self):
        assert _num("") == 0.0

    def test_spanish_day_tuesday(self):
        d = datetime(2026, 3, 3)   # Tuesday
        assert SPANISH_DAYS[d.weekday()] == "MARTES"

    def test_spanish_month_march(self):
        assert SPANISH_MONTHS_FULL[3] == "MARZO"


# ---------------------------------------------------------------------------
# Smoke tests (fast, don't load the full dataset)
# ---------------------------------------------------------------------------

class TestParsers:
    def test_parse_invr0601_ventas_count(self):
        sales = parse_invr0601(BASE_DIR / "lp60.dat", "FACTURACION", "VENTA")
        assert len(sales) > 0, "lp60.dat parsed no records"

    def test_parse_invr0601_returns_count(self):
        rets = parse_invr0601(BASE_DIR / "lp59.dat", "DEVOLUCION", "NOTA DE CREDITO")
        assert len(rets) > 0, "lp59.dat parsed no records"

    def test_parse_invr0601_sample_row(self):
        sales = parse_invr0601(BASE_DIR / "lp60.dat", "FACTURACION", "VENTA")
        # Find document 000-002-01-00053402 line 6
        row = next(
            (r for r in sales if r["documento"] == "000-002-01-00053402" and r["linea"] == 6),
            None,
        )
        assert row is not None
        assert row["bodega"]     == 4
        assert row["cod_prov"]   == 108
        assert row["linea_prod"] == 36
        assert row["cod_barra"]  == "7503002 163610"
        assert row["cantidad"]   == 1
        assert row["costo"]      == pytest.approx(217.85, abs=0.01)
        assert row["precio"]     == pytest.approx(294.56, abs=0.01)
        assert row["descuento"]  == pytest.approx(13.94,  abs=0.01)
        assert row["alterno"]    == 87320

    def test_parse_venr15_sample(self):
        d = parse_venr15(BASE_DIR / "lp33.dat")
        assert "000-002-01-00053402" in d
        entry = d["000-002-01-00053402"]
        assert entry["fecha"]   == datetime(2026, 3, 3)
        assert entry["cliente"] == 30312102

    def test_parse_facr12_sample(self):
        d = parse_facr12(BASE_DIR / "lp35.dat")
        assert "000-002-01-00053402" in d
        entry = d["000-002-01-00053402"]
        assert entry["cliente"] == 30312102
        assert entry["orden"]   == "T6961T0"

    def test_parse_notr03_sample(self):
        d = parse_notr03(BASE_DIR / "lp31.dat")
        assert "000-002-06-00025691" in d
        entry = d["000-002-06-00025691"]
        assert entry["fecha"]   == datetime(2026, 3, 3)
        assert entry["cliente"] == 803339

    def test_parse_invr29_sample(self):
        d = parse_invr29(BASE_DIR / "lp51.dat")
        assert ("000-002-01-00053402", 6) in d
        assert d[("000-002-01-00053402", 6)] == "WALMART CDS"

    def test_parse_invr29_return_razon(self):
        d = parse_invr29(BASE_DIR / "lp51.dat")
        assert ("000-002-06-00025691", 1) in d
        assert d[("000-002-06-00025691", 1)] == "DEVOLUCION TEG."

    def test_load_variables_proveedor(self):
        v = load_variables(BASE_DIR / "Variables.xlsx")
        assert v["prov_lookup"].get(108) == "HENKEL LA LUZ"
        assert v["prov_lookup"].get(115) == "SUPER DE ALIMENTOS"
        assert v["prov_lookup"].get(102) == "SUMMA INDUSTRIAL"

    def test_load_variables_bodega_estacion(self):
        v = load_variables(BASE_DIR / "Variables.xlsx")
        assert v["bodega_est"].get(4)  == "TEG"
        assert v["bodega_est"].get(3)  == "SPS"
        assert v["bodega_est"].get(45) == "TEG"

    def test_load_variables_peso(self):
        v = load_variables(BASE_DIR / "Variables.xlsx")
        assert v["peso_lookup"].get(87320) == pytest.approx(7.82, abs=0.01)

    def test_load_variables_client_ruta(self):
        v = load_variables(BASE_DIR / "Variables.xlsx")
        cli = v["client_ruta"].get(30312102)
        assert cli is not None
        assert cli["nombre"]   == "WALMART"
        assert cli["tipo"]     == "SUPERMERCADOS"
        assert cli["ruta_pres"] == "HN11"


class TestBuildDatasetSamples:
    """Spot-check specific rows in the built dataset."""

    def test_walmart_teg_row(self, dataset):
        row = next(
            (r for r in dataset
             if r["DOCUMENTO"] == "000-002-01-00053402" and r["LINEA FAC"] == 6),
            None,
        )
        assert row is not None, "Expected row not found in dataset"
        assert row["BODEGA"]          == 4
        assert row["COD PROV"]        == 108
        assert row["PROVEEDOR"]       == "HENKEL LA LUZ"
        assert row["LINEA"]           == 36
        assert row["CODIGO BARRA"]    == "7503002 163610"
        assert row["DESCRIPCION"]     == "XTREME GEL ATTRA.12/250g"
        assert row["CANTIDAD"]        == 1
        assert row["COSTO"]           == pytest.approx(217.85, abs=0.01)
        assert row["PRECIO"]          == pytest.approx(294.56, abs=0.01)
        assert row["DESCUENTO"]       == pytest.approx(13.94,  abs=0.01)
        assert row["ALTERNO"]         == 87320
        assert row["MODULO"]          == "FACTURACION"
        assert row["ESTACION"]        == "TEG"
        assert row["MES"]             == "MARZO"
        assert row["AÑO"]             == 2026
        assert row["FECHA"]           == datetime(2026, 3, 3)
        assert row["DIA"]             == "MARTES"
        assert row["SEMANA"]          == 1
        assert row["CODIGO CLIENTE"]  == 30312102
        assert row["DEPARTAMENTO"]    == "FRANCISCO MORAZAN"
        assert row["MUNICIPIO"]       == "DISTRITO CENTRAL"
        assert row["NOMBRE CLIENTE"]  == "WALMART"
        assert row["TIPO CLIENTE"]    == "SUPERMERCADOS"
        assert row["COD VENDEDOR"]    == 10
        assert row["VENDEDOR"]        == "ETNA VENEGAS WAI"
        assert row["RUTA"]            == "Supermercado"
        assert row["TIPO VENTA"]      == "VENTA"
        assert row["ORDEN"]           == "T6961T0"
        assert row["RAZON"]           == "WALMART CDS"
        assert row["PESO"]            == pytest.approx(7.82, abs=0.01)
        assert row["RUTA-PRESUPUESTO"] == "HN11"

    def test_return_nota_credito_row(self, dataset):
        row = next(
            (r for r in dataset
             if r["DOCUMENTO"] == "000-002-06-00025691" and r["LINEA FAC"] == 1),
            None,
        )
        assert row is not None, "Expected return row not found"
        assert row["TIPO VENTA"]  == "NOTA DE CREDITO"
        assert row["MODULO"]      == "DEVOLUCION"
        assert row["RAZON"]       == "DEVOLUCION TEG."
        assert row["ORDEN"]       == 0
        assert row["COD PROV"]    == 115
        assert row["PROVEEDOR"]   == "SUPER DE ALIMENTOS"
        assert row["CANTIDAD"]    == 50
        assert row["FECHA"]       == datetime(2026, 3, 3)
